from zipfile import ZipFile
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
from paths import *

def test_read_csv():
    with ZipFile(archive_path, 'r') as zf:
        archive = zf.read('Csv.csv').decode()
        default = open(os.path.join(resources_path, 'Csv.csv')).read()
        assert archive == default

def test_read_xlsx():
    with ZipFile(archive_path, 'r') as zf:
        archive = load_workbook(zf.open('Excel.xlsx'))
        default = load_workbook(os.path.join(resources_path, 'Excel.xlsx'))

        sheet_archive = archive.active
        sheet_default = default.active

        archive_value = []
        default_value = []

        for i in range(0, sheet_archive.max_row):
            for j in sheet_archive.iter_cols(1, sheet_archive.max_column):
                archive_value.append(j[i].value)

        for i in range(0, sheet_default.max_row):
            for j in sheet_default.iter_cols(1, sheet_default.max_column):
                default_value.append(j[i].value)

        assert archive_value == default_value
        assert default.sheetnames == archive.sheetnames

def test_read_pdf():
    with ZipFile(archive_path, 'r') as zf:
        archive = PdfReader(zf.open('Pdf.pdf'))
        default = PdfReader(os.path.join(resources_path, 'Pdf.pdf'))
        assert len(archive.pages) == len(default.pages)
        assert archive.pages[0].extract_text() == default.pages[0].extract_text()
