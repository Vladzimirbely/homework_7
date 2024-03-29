from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
from conftest import archive
import csv

def test_read_csv():
    with ZipFile(archive) as zip_file:
        with zip_file.open('Csv.csv') as csv_file:
            content = csv_file.read().decode()
            csvreader = list(csv.reader(content.splitlines()))
            row = csvreader[0]

            assert row == ['Username; Identifier;First name;Last name']

def test_read_pdf():
    with ZipFile(archive) as zip_file:
        reader = PdfReader(zip_file.open('Pdf.pdf'))
        number_of_pages = len(reader.pages)
        page = reader.pages[0]
        text = page.extract_text()

        assert number_of_pages == 1
        assert text == 'Пример pdf '

def test_read_xlsx():
    with ZipFile(archive, 'r') as zip_file:
        workbook = load_workbook(zip_file.open('Excel.xlsx'))
        sheet = workbook.active

        assert sheet.cell(row=2, column=2).value == 'Hello'

        for cols in sheet.iter_cols():
            assert len(cols) == 5

        for rows in sheet.iter_rows():
            assert len(rows) == 2
